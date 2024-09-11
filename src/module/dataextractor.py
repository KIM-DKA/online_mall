import pandas as pd
from openpyxl import load_workbook
from pyxlsb import open_workbook

class Extractor:
    
    @classmethod
    def data_load(cls, path: str, header: str=None, sheet_list: list=None):
        """
        workbook 데이터 로드하는 함수
        
        :param path : 가져올 데이터의 경로
        """
        if path.endswith(('xls', 'xlsx', 'xlsm')):
            return cls._get_data_for_standard(path, header, sheet_list)
        
        elif path.endswith(('xlsb')):
            return cls._get_data_for_binary(path, header, sheet_list)
        
    @classmethod
    def _get_data_for_standard(cls, path, header, sheet_list):
        data_dict = dict()
        workbook = load_workbook(path, read_only=True, data_only=True)
        
        for sheet in workbook.sheetnames:
            
            if sheet_list:
                if sheet not in sheet_list:
                    continue
            
            data = workbook[sheet].values
            
            if not header:
                sheet_table = pd.DataFrame(data).copy()
                data_dict[sheet] = sheet_table
                continue
            
            columns = next(data)
            while header not in columns:
                columns = next(data)
                
            sheet_table = pd.DataFrame(data, columns=columns).copy()
            try:
                sheet_table.rename(columns=lambda col: col.strip(), inplace=True)
            except:
                pass
            data_dict[sheet] = sheet_table
            
        return data_dict
    
    @classmethod
    def _get_data_for_binary(cls, path, header, sheet_list):
        data_dict = dict()
        with open_workbook(path) as wb:
            for sheet in wb.sheets:
                
                if sheet_list:
                    if sheet not in sheet_list:
                        continue
                
                rows = []
                with wb.get_sheet(sheet) as sh:
                    for row in sh.rows():
                        row_value = [item.v for item in row]
                        if header in row_value:
                            columns = row_value
                            continue
                        rows.append(row_value)
                        
                data_dict[sheet] = pd.DataFrame(rows, columns=columns)
        return data_dict